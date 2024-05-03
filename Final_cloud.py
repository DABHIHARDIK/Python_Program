import openpyxl
from openpyxl import load_workbook
excel_file ="C:\\Users\\nxg02041\\OneDrive - NXP\\Documents\\C_program\\New folder\\ADVANCE _Wi-Fi.xlsx"
wb = load_workbook(excel_file)
grey = 23 
n=1
m=0
check=0 
output=""
precolor=grey
curcolor=0
colors=[]
numbers=[]
greyindex = 0
active_grey=0
sheetname = wb.sheetnames
for l in range (2,len(sheetname)-3):
    sh = wb[sheetname[l]]
    print(sheetname[l])
    active_grey=0
    greyindex = 0
    raw=sh.max_row
    print("max row = ",raw)
    for i in range (1,raw):
        colorfound=0
        output=""
        output="_"
        istring = str(i) 
        index= 'AC' + istring 
        color_in_hex = sh[index].fill.start_color.index # this gives you Hexadecimal value of the color
        if(color_in_hex == grey):
            if(active_grey==0):
                numbers.append(1)
                colors.append(color_in_hex)
            else:
                numbers[0]+=1
            greyindex += 1 
            strgreyindex = '_'+ str(greyindex)
            mycell = sh [index]
            mycell.value = strgreyindex
            active_grey+=1
            for h in range(1,len(numbers)):
                numbers[h]=0
            continue
        if(active_grey>=1):
            colorfound=0
            mycell = sh [index]
            curcolor=color_in_hex
            if(precolor!=curcolor):
                check=0
                for j in range(0,len(colors)):
                    if(colors[j]==curcolor):
                        colorfound=j
                        numbers[j]+=1
                        for one in range(j+1,len(numbers)):
                            numbers[one]=0
                        continue
                        
                if(colorfound==0):
                    colors.append(curcolor)
                    numbers.append(0)
                    n+=1
                    for m in range(0,len(colors)):
                        if(colors[m]==curcolor):
                            numbers[m]+=1
                    #print("1")
                    for w in range(0,len(numbers)):
                        if(check==0):
                            output+=str(numbers[w]) 
                            check+=1
                        else:
                            output+='.' + str(numbers[w])               
                    mycell.value = output
                    #print(output)
                    
                else:
                    #print("2")
                    for k in range(0,colorfound+1):
                        
                        if(check==0):
                            output+=str(numbers[k]) 
                            check+=1
                        else:
                            output+='.' + str(numbers[k])               
                    mycell.value = output

                precolor=curcolor
                #print(output)
            else:
                check=0
                #print("3")
                for m in range(0,len(colors)):
                    if(colors[m]==curcolor):
                        numbers[m]+=1
                        continue
                for k in range(0,len(numbers)):
                    
                    if(check==0):
                        output+=str(numbers[k])
                        check+=1
                    else:
                        output+='.' + str(numbers[k])               
               
                mycell.value = output
                precolor=curcolor
                #print(output)
    colors=[]
    numbers=[]
    print(colors)
    print(numbers)
wb.save(excel_file)
